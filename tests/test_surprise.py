import sys
from hhmcds.model import 早安鬧鐘資料
from hhmcds.surprise import 抽獎
from hhmcds.smurf import 分身
from hhmcds.imei import get_imei
import datetime
from openpyxl import workbook, styles
from hhmcds.gmail import active_then_delete
import time


def 抽獎流程(gmail, password, mail_prefix, numbers):
    wb = workbook.Workbook()
    ws = wb.active
    ws.title = "抽獎結果"
    ws.freeze_panes = 'A2'
    ws.cell(row=1, column=1).value = "信箱"
    ws.cell(row=1, column=1).font = styles.Font(bold=True)
    ws.cell(row=1, column=2).value = "結果"
    ws.cell(row=1, column=2).font = styles.Font(bold=True)

    count = 1
    str_imei = get_imei()
    for idx in range(1, numbers):
        資料 = 早安鬧鐘資料("{0}+{1:}{2:03d}@gmail.com".format(gmail, mail_prefix, idx), password, str_imei)
        抽獎物件 = 抽獎(資料)
        回傳資料 = 抽獎物件.執行()
        if 回傳資料.抽獎成功:
            count += 1
            ws.cell(row=count, column=1).value = 回傳資料.信箱
            ws.cell(row=count, column=2).value = 回傳資料.抽獎結果
            sys.stdout.write("程序執行完成\n\n")
        else:
            sys.stderr.write("中斷\n")

    wb.save(filename="抽獎結果{}-{:%m%d}.xlsx".format(mail_prefix, datetime.datetime.now()))


def 創建分身(gmail, password, mail_prefix, numbers):
    count = 0
    str_imei = get_imei()

    for idx in range(1, numbers):
        count += 1
        if count % 10 is 0:
            str_imei = get_imei()
        資料 = 早安鬧鐘資料("{0}+{1:}{2:03d}@gmail.com".format(gmail, mail_prefix, idx), password, str_imei)
        分身物件 = 分身(資料)

        if 分身物件.執行():
            sys.stdout.write("程序執行完成\n\n")
        else:
            sys.stderr.write("中斷\n")


def 測試程式(gmail, password, mail_prefix="", numbers=100):
    #創建分身(gmail, password, mail_prefix, numbers)
    #sys.stdout.write("分身創建完成，等待五分鐘後開始繳活\n\n")
    #time.sleep(600)
    active_then_delete()
    抽獎流程(gmail, password, mail_prefix, numbers)

if __name__ == "__main__":
    mail_prefix = "{:%m%d}".format(datetime.datetime.now())
    測試程式(gmail="HappyHackingNinja", password="md1234", mail_prefix=mail_prefix, numbers=200)
